def num_there(s):
    return any(i.isdigit() for i in s)

import re
import os
import time
import math

from datetime import datetime
from dateutil.relativedelta import relativedelta
import pickle 
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By


from openpyxl import load_workbook

def containsNumber(value):
    return any([char.isdigit() for char in value])

# The profile where I enabled the VPN previously using the GUI.
opera_profile = r'Some Path' 
options = webdriver.ChromeOptions()
options.add_argument('user-data-dir=' + opera_profile)
options._binary_location = r'Some Path'

minn=2
maxn=113
maxins=5

from selenium import webdriver
from webdriver_manager.opera import OperaDriverManager

driver = webdriver.Opera(executable_path=OperaDriverManager().install())
quit()

url_0="Some URL"
url="Some URL"
url_2="Some URL 2"

maxrep=int(math.ceil((maxn-minn)/maxins))
print(maxrep)
for j in range(0,maxrep):

    os.system("taskkill /im opera.exe /f")
    
    driver1 = webdriver.Opera(options=options)
    time.sleep(5)
    driver1.get(url_0)
    time.sleep(10)
    try:
       driver1.find_element_by_xpath("//*[contains(text(), 'Вход')]").click()
    except:
       pass 
    if driver1.current_url==url:
        time.sleep(2)
        login=driver1.find_element_by_name("UserLogin")
        login.send_keys("")
        password=driver1.find_element_by_name("UserPassword")
        password.send_keys("")
        driver1.find_element_by_name("LoginBtn").click()
    
    break
    
    c = driver1.get_cookies()
    pickle.dump( driver1.get_cookies() , open("cookies.pkl","wb"))   
    #driver1.maximize_window()
    driver1.get(url_2)
    
    wb = load_workbook("")
    res_wb=wb
    ws1=wb.active
    
    path=""
    ref_workbook = load_workbook(path)
    ws=ref_workbook.active
    
    driver1.find_element_by_xpath("//*[contains(text(), 'Some Text')]").click()
    time.sleep(2)
    driver1.find_element_by_xpath("//*[contains(text(), 'Some Text')]").click()
    time.sleep(2)
    
    clearlist=driver1.find_elements_by_xpath("//*[contains(text(), 'Some Text')]")
    for elems in clearlist:
           elems.click()
           time.sleep(2)

    fedbtn = driver1.find_element_by_xpath("//*[contains(text(),'Some Text')]")
    fedbtnanc=fedbtn.find_element_by_xpath('..') 
    fedbtnanc.find_element_by_xpath(".//input[1]").click() 
    fedbtn = driver1.find_element_by_xpath("//*[contains(text(),'Some Text')]")
    fedbtnanc=fedbtn.find_element_by_xpath('..') 
    fedbtnanc.find_element_by_xpath(".//input[1]").click() 


    myElem = WebDriverWait(driver1, 10).until(EC.presence_of_element_located((By.ID, "Some HTML")))

    for x in range (0, maxins):
     
     rownum=(x)+maxins*j+minn
    
     name2=ws.cell(row=rownum,column=16).value
     name='("и") или (в) или ("не") или (на) или (я) или (быть) или (он) или (с) или (что) или (а) или (по) или (это) или (она) или (этот) или (к) или (но) или (они) или (мы) или (как) или (из) или (у) или (который) или (то) или (за) или (свой) или (что) или (весь) или (год) или (от) или (так) или (о) или (для) или (ты) или (же) или (все) или (тот) или (мочь) или (вы) или (человек) или (такой) или (его) или (сказать) или (только) или ("или") или (ещё) или (бы) или (себя) или (один) или (как) или (уже) или (до) или (время) или (если) или (сам) или (когда) или (другой) или (вот) или (говорить) или (наш) или (мой) или (знать) или (стать) или (при) или (чтобы) или (дело) или (жизнь) или (кто) или (первый) или (очень) или (два) или (день) или (её) или (новый) или (рука) или (даже) или (во) или (со) или (раз) или (где) или (там) или (под) или (можно) или (ну) или (какой) или (после) или (их) или (работа) или (без) или (самый) или (потом) или (надо) или (хотеть) или (ли) или (слово) или (идти) или (большой) или (должен) или (место) или (иметь) или (ничто)'
     region=ws.cell(row=rownum,column=1).value
     startdate=ws.cell(rownum,5).value
     enddate=ws.cell(rownum,6).value
     
     print(name, startdate, enddate)

      
     startdateqs=datetime.strptime(startdate, '%d/%m/%Y')
     startdate2qs=startdateqs+relativedelta(months=1)
     enddateqs=datetime.strptime(enddate, '%d/%m/%Y')
      
     mindate=datetime(2000, 1, 1)
     if startdateqs<mindate:
       startdateqs=mindate 
     maxdate=datetime(2012,4,1)
     if enddateqs>maxdate:
         enddateqs=maxdate
     nstartmonth = (startdateqs.year - 2000) * 12 + (startdateqs.month - 1)   
     nendmonth = (enddateqs.year - 2000) * 12 + (enddateqs.month - 1)  
    
     name2a=name
     list_num_original=[]
     list_num = []
     res_num=[]
     res_list=[]
     date_list=[]
     searchbox=driver1.find_element_by_name("Some HTML")
     searchbox.clear()
     searchbox.send_keys(name2a)
     dat=startdateqs
     
     while dat<enddateqs+relativedelta(months=1):  
     
      date1_time = dat.strftime("%Y-%m-%d %H:%M:%S")    
      date1_time_check=dat.strftime("%Y-%m-%d")
     
      datstartwindow=driver1.find_element_by_id("Some HTML")
      driver1.execute_script("arguments[0].setAttribute('value','"+date1_time+"')", datstartwindow)
     
      dat=dat+relativedelta(months=1)
      date2_time = dat.strftime("%Y-%m-%d %H:%M:%S")
      date2_time_check=dat.strftime("%Y-%m-%d")
        
      datendwindow= driver1.find_element_by_id("Some HTML")
      driver1.execute_script("arguments[0].setAttribute('value','"+date2_time+"')", datendwindow)
     
      myElem=WebDriverWait(driver1,50).until(EC.presence_of_element_located((By.CSS_SELECTOR, "input[value='"+date1_time_check+"']")))
      myElem=WebDriverWait(driver1,50).until(EC.presence_of_element_located((By.CSS_SELECTOR, "input[value='"+date2_time_check+"']")))
     
      searchbttn=driver1.find_element_by_id("Some HTML")
      searchbttn.click()
      
      attempt=0
      res_num.clear
      res=""
      
      while attempt<6:
       try:
         element = WebDriverWait(driver1, 50).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#Some HTML")))
         attempt=10
       except:
        attempt=attempt+1
        searchbttn.click()
      res=driver1.find_element_by_id("Some HTML").text

      res_num=[int(s) for s in re.findall(r'\d+',res)]
      date_list.append(dat)
      res_list.append(res)
      #print(res_num)
      attempt=0
      if containsNumber(res)==False:
          res_num.append("-")
      while attempt<3:
       try:
        list_num.append(res_num[0])
        attempt=10
       except:
        res_wb.save("Some Path")
        searchbttn=driver1.find_element_by_id("Some HTML")  
        searchbttn.click()
        attempt=attempt+1
       
      attempt=0
 
     for i in range(nstartmonth, nendmonth+1):
          ws1.cell(i+144*(rownum-1),80).value= list_num[i-nstartmonth]
     
     print(x)
     res_wb.save("Some Path")